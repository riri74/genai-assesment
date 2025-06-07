import os
import pandas as pd
from openpyxl import load_workbook
import requests
import json
import time
import difflib

def call_groq(prompt, retries=3, backoff_factor=1):
    api_key = os.getenv("GROQ_API_KEY")
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "model": "llama3-8b-8192",
        "messages": [
            {"role": "system", "content": "You are a smart assistant that maps data labels to their closest matches."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.2
    }

    for i in range(retries):
        response = requests.post(url, headers=headers, data=json.dumps(data))
        if response.status_code == 429:
            wait = backoff_factor * (i + 1)
            print(f"Rate limit hit. Waiting {wait} seconds before retrying...")
            time.sleep(wait)
        else:
            response.raise_for_status()
            return response.json()["choices"][0]["message"]["content"].strip()

    raise Exception("Rate limit exceeded after retries.")

def match_placeholder_with_ai(placeholder, data_summary):
    prompt = f"""
Given this placeholder from an Excel template: "{placeholder}"

And this summary of available source data fields:

{data_summary}

Which field (or role) from the source data most likely corresponds to the placeholder?

Respond with just the exact field name or role, nothing else.
"""
    return call_groq(prompt)

def is_semantic_match(placeholder, matched_key):
    placeholder_lower = placeholder.lower()
    matched_key_lower = matched_key.lower()

    if any(word in placeholder_lower for word in ["nurse", "care worker", "care minutes", "personal care", "staff", "management", "allied health"]):
        if any(bad_word in matched_key_lower for bad_word in ["bedday", "occupiedbeddays", "availablebeddays"]):
            return False
    if "bed day" in placeholder_lower:
        if not any(good_word in matched_key_lower for good_word in ["bedday", "occupiedbeddays", "availablebeddays"]):
            return False
    if "rate" in placeholder_lower:
        if "rate" not in matched_key_lower:
            return False

    return True

def fallback_match(placeholder, keys):
    placeholder_clean = placeholder.strip("◦ ").lower()
    keys_lower = [k.lower() for k in keys]
    matches = difflib.get_close_matches(placeholder_clean, keys_lower, n=1, cutoff=0.4)
    if matches:
        index = keys_lower.index(matches[0])
        return keys[index]
    return None

def build_aggregated_costs(source_paths):
    print("Aggregating data from source files...")
    combined = {}
    for path in source_paths:
        df = pd.read_csv(path)
        df = df.select_dtypes(include='number')
        if 'Role' in df.columns and 'Cost_AUD' in df.columns:
            combined.update(df.groupby('Role')['Cost_AUD'].sum().to_dict())
        elif 'Field' in df.columns and 'Value' in df.columns:
            combined.update(df.groupby('Field')['Value'].sum().to_dict())
        else:
            for col in df.columns:
                combined[col] = combined.get(col, 0) + df[col].sum()
    return combined

def populate_template_excel(template_path, output_path, source_paths):
    wb = load_workbook(template_path)
    ws = wb.active

    cost_lookup = build_aggregated_costs(source_paths)
    summary_str = "\n".join(f"{k}: {v}" for k, v in cost_lookup.items())
    keys = list(cost_lookup.keys())

    print("Populating template using Groq...")

    mappings = {}
    success = 0
    total = 0
    suspicious_count = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().startswith("◦"):
                placeholder = cell.value.strip()
                total += 1
                try:
                    matched_key = match_placeholder_with_ai(placeholder, summary_str)

                    if not is_semantic_match(placeholder, matched_key):
                        fallback_key = fallback_match(placeholder, keys)
                        if fallback_key and is_semantic_match(placeholder, fallback_key):
                            matched_key = fallback_key
                        else:
                            print(f"Rejected AI match '{matched_key}' for placeholder '{placeholder}', no valid fallback found.")
                            suspicious_count += 1
                            continue

                    mappings[placeholder] = matched_key
                    value = cost_lookup.get(matched_key)
                    if value is not None:
                        ws.cell(row=cell.row, column=cell.column + 1).value = round(value, 2)
                        success += 1
                except Exception as e:
                    print(f"Failed for placeholder '{placeholder}': {e}")
                    suspicious_count += 1

    wb.save(output_path)
    print(f"Template saved to: {output_path}")

    print("\nFinal Mappings:")
    for k, v in mappings.items():
        print(f"→ {k} → {v}")

    accuracy = (success / total) * 100 if total > 0 else 0
    data_correctness = ((success - suspicious_count) / success) * 100 if success > 0 else 0

    print(f"\nGroq Matching Accuracy: {accuracy:.2f}%")
    print(f"Data Correctness: {data_correctness:.2f}%")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))

    template_file = os.path.join(base_dir, "Template File", "template.xlsx")
    output_file = os.path.join(base_dir, "output_filled_template.xlsx")
    source_files = [
        os.path.join(base_dir, "Source File", "agency_staff_costs.csv"),
        os.path.join(base_dir, "Source File", "employee_labour_costs.csv"),
        os.path.join(base_dir, "Source File", "bed_days.csv"),
        os.path.join(base_dir, "Source File", "labour_hours.csv"),
        os.path.join(base_dir, "Source File", "hourly_rates.csv"),
        os.path.join(base_dir, "Source File", "outbreak_management_costs.csv")
    ]

    populate_template_excel(template_file, output_file, source_files)
